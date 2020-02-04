from collections import Counter
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles.colors import WHITE
from openpyxl import Workbook
import datetime
import smtplib
from email import encoders
from os import listdir
from os.path import isfile, join

def lista_facturas_credito(ws):
    x = 2
    list_facturas_cred = []
    while ws['B' + str(x)].value:
      if ws['E' + str(x)].value == 'INTERESES':
        if ws['G' + str(x)].value > 0:
            list_facturas_cred.append(int(ws['B' + str(x)].value))
      x += 1
    list1 = list(set(list_facturas_cred))
    return list1

def lectura_reporte(ws,list_fact):
    x = 2
    list = []
    while ws['B' + str(x)].value:
      if ws['B' + str(x)].value in list_fact:
          if ws['F' + str(x)].value != 'BOLSAS Y TERMOENCOGI':
              if ws['F' + str(x)].value != 'EQUIPOS COMPUTO TIEN':
                 if ws['E' + str(x)].value == 'INTERESES':
                    list.append([int(ws['B' + str(x)].value),ws['C' + str(x)].value,int(ws['G' + str(x)].value),int(ws['H' + str(x)].value),ws['O' + str(x)].value,"int", int(ws['J' + str(x)].value)])
                 else:
                    list.append([int(ws['B' + str(x)].value),ws['C' + str(x)].value,int(ws['G' + str(x)].value),int(ws['H' + str(x)].value),ws['O' + str(x)].value,"None",int(ws['J' + str(x)].value)])

      x += 1
    #list No factura, fecha, cantidad, valor unitario, vendedor
    return list

def crear_db(list):
    fechas = {}
    facturas1 = {}
    vendedor = {}
    vendedor1_fact = {}
    vendedor_fecha = {}
    vendedor2 = {}
    lista_por_fecha = []

    for l in list:
        fechas.update({l[1] : None})
        facturas1.update({l[0] : None})
        vendedor.update({l[4] : None})
        vendedor1_fact.update({l[4] : None})

    #Crear el diccionario de facturas totales
    for k,v in vendedor.items():
        list2 = []
        for l2 in list:
            if k == l2[4]:
                list2.append([l2[2],l2[3],l2[0],l2[5],l2[6]])
        vendedor.update({k : list2})


    for k1,v1 in vendedor1_fact.items():
        list3 = []
        for l3 in list:
            if k1 == l3[4]:
                list3.append([l3[2],l3[3],l3[0],l3[1],l3[5],l3[6]])
        vendedor_fecha.update({k1 : list3})

    for k,v in vendedor_fecha.items():
        for k1, v1 in fechas.items():
            venta_dia = 0
            for l in v:
                if k1 == l[3]:
                    if l[4] == 'int':
                        venta_dia = venta_dia + (int(l[1]) * int(l[0]))
                    else:
                        if int(l[5]) > 0:
                            venta_dia = venta_dia + round(((int(l[1]) * int(l[0])) * 1.19) * (1 - float(l[5])/100))
                        else:
                            venta_dia = venta_dia + round(((int(l[1]) * int(l[0])) * 1.19))
            lista_por_fecha.append([k, k1, venta_dia])
    return vendedor, lista_por_fecha

def procesar_db(vendedor):
    list_vendedora_fecha = []
    list_vendedora_total = []
    list_por_vendedora = []
    venta_dia = 0
    venta_total = 0

    #crear lista con vendedor, venta total reporte
    for k,v in vendedor.items():
        venta_total = 0
        for l1 in v:
            if l1[3] == "int":
                venta_total = venta_total + (int(l1[0]) * int(l1[1]))
            else:
                if l1[4] > 0:
                    venta_total = venta_total + round(((int(l1[0]) * int(l1[1])) * 1.19) * (1 - float(l1[4])/100))
                else:
                    venta_total = venta_total + round(((int(l1[0]) * int(l1[1])) * 1.19))
        list_vendedora_total.append([k, venta_total])
    return list_vendedora_total


def print_report(list_vendedora_total, lista_por_fecha, file):
    excel_salida = openpyxl.Workbook()


# formato reporte
    blackFill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
    lightgreyFill = PatternFill(start_color='E0E0E0E0', end_color='E0E0E0E0', fill_type='solid')
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')


    excel_salida.create_sheet(index=0, title='Informe de Ventas a Credito')
    hoja_salida = excel_salida.worksheets[0]
    excel_salida.create_sheet(index=1, title='Ventas a credito por vendedora')
    hoja_salida1 = excel_salida.worksheets[1]

    hoja_salida['A1'] = 'Vendedora'
    hoja_salida['B1'] = 'Venta credito total'
    hoja_salida1['A1'] = 'Vendedora'
    hoja_salida1['B1'] = 'Fecha'
    hoja_salida1['C1'] = 'Ventas a credito'

    for cell in 'ABC':
        hoja_salida[cell +'1'].fill = blackFill
        hoja_salida[cell +'1'].font = Font(color=WHITE, bold=True, size=20)
        hoja_salida1[cell +'1'].fill = blackFill
        hoja_salida1[cell +'1'].font = Font(color=WHITE, bold=True, size=20)


    dim = {'A': 40, 'B': 35, 'C': 20}
    for k, v in dim.items():
        hoja_salida.column_dimensions[k].width = v
        hoja_salida1.column_dimensions[k].width = v

# Escribir Totales
    row = 2
    for i in list_vendedora_total:
        hoja_salida['A' + str(row)] = i[0]
        hoja_salida['B' + str(row)] = i[1]
        row += 1

#Escribir ventas por fecha
    row1 = 2
    for i in lista_por_fecha:
        hoja_salida1['A' + str(row1)] = i[0]
        hoja_salida1['B' + str(row1)] = i[1]
        hoja_salida1['C' + str(row1)] = i[2]
        row1 += 1

    fecha = datetime.date.today()
    excel_salida.save("/Users/mpatinob/Dropbox/Pruebas/Reportes Procesados/"+ str(fecha) + "_Procesado_Cred_" + str(file))


path = '/Users/mpatinob/Dropbox/Pruebas/Reportes SAP'
files = [f for f in listdir(path) if isfile(join(path, f))]

for file in files:
    if file != '.DS_Store':
        try:
            wb = openpyxl.load_workbook('/Users/mpatinob/Dropbox/Pruebas/Reportes SAP/' + str(file))
            ws = wb.worksheets[0]
            lista_fact = lista_facturas_credito(ws)
            lista_inicial = lectura_reporte(ws,lista_fact)
            vendedor, lista_por_fecha = crear_db(lista_inicial)
            list_vendedora_total = procesar_db(vendedor)
            print_report(list_vendedora_total, lista_por_fecha, file)
        except KeyError:
            print "Archivo Corrupto o No soportado"
