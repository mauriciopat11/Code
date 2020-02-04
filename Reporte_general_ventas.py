from collections import Counter
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.styles.colors import WHITE
from openpyxl import Workbook
import datetime
import smtplib
from email import encoders
from os import listdir
from os.path import isfile, join

def lectura_reporte(ws):
    x = 2
    list = []
    while ws['B' + str(x)].value:
      if ws['F' + str(x)].value != 'BOLSAS Y TERMOENCOGI':
            if ws['F' + str(x)].value != 'EQUIPOS COMPUTO TIEN':
                if ws['E' + str(x)].value != 'INTERESES':
                    if ws['G' + str(x)].value > 0:
                        list.append([int(ws['B' + str(x)].value),ws['C' + str(x)].value,int(ws['G' + str(x)].value),ws['O' + str(x)].value])
      x += 1
    return list


def crear_db(list):
    fechas = {}
    facturas1 = {}
    vendedor = {}
    for l in list:
        fechas.update({l[1] : None})
        facturas1.update({l[0] : None})
        vendedor.update({l[3] : None})

    for k,v in facturas1.items():
        list1 = []
        for l1 in list:
            if k == l1[0]:
                list1.append([l1[2],l1[3]])
        facturas1.update({ k : list1})

    for k,v in fechas.items():
        list1 = []
        facturas = {}
        for l1 in list:
            if k == l1[1]:
                facturas.update({l1[0] : facturas1[l1[0]]})
        fechas.update({ k : facturas})
    return fechas, vendedor

def procesar_db(db):
    list_total_reporte = []
    list_prendas_facturas_dia = []
    list_por_vendedora = []
    conteo_facturas = 0
    conteo_prendas = 0

    for k,v in db.items():
        conteo_facturas = conteo_facturas + len(v)
        for k1,v1 in v.items():
            for l in v1:
                conteo_prendas = conteo_prendas + l[0]

    list_total_reporte.append([conteo_facturas , conteo_prendas])


    for k,v in db.items():
        conteo_prendas_dia = 0
        for k1,v1 in v.items():
            for l in v1:
                conteo_prendas_dia = conteo_prendas_dia + l[0]
        list_prendas_facturas_dia.append([k, len(v) , conteo_prendas_dia])

    #fecha, numero de facturas, numero de prendas


    for k,v in db.items():
        for k2,v2 in vendedor.items():
            conteo_vendedora_prendas = 0
            conteo_vendedora_facturas = 0
            for k1,v1 in v.items():
                if k2 == v1[0][1]:
                    conteo_vendedora_facturas = conteo_vendedora_facturas + 1
                for l in v1:
                    if k2 == l[1]:
                        conteo_vendedora_prendas = conteo_vendedora_prendas + l[0]
            list_por_vendedora.append([k, k2, conteo_vendedora_facturas, conteo_vendedora_prendas])

    return list_total_reporte, list_prendas_facturas_dia, list_por_vendedora

def print_report(list_total_reporte, list_prendas_facturas_dia, list_por_vendedora, file):
    excel_salida = openpyxl.Workbook()


# formato reporte
    blackFill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
    lightgreyFill = PatternFill(start_color='E0E0E0E0', end_color='E0E0E0E0', fill_type='solid')
    redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')


    excel_salida.create_sheet(index=0, title='Informe de Ventas General')
    hoja_salida = excel_salida.worksheets[0]
    excel_salida.create_sheet(index=1, title='Informe de Ventas por Vendedora')
    hoja_salida1 = excel_salida.worksheets[1]

    hoja_salida['A1'] = 'Total Facturas Reporte'
    hoja_salida['B1'] = 'Total Prendas Reporte'
    hoja_salida['A5'] = 'Fecha'
    hoja_salida['B5'] = '# Facturas'
    hoja_salida['C5'] = '# Prendas'
    hoja_salida1['A1'] = 'Fecha'
    hoja_salida1['B1'] = '# Facturas'
    hoja_salida1['C1'] = '# Prendas'
    hoja_salida1['D1'] = 'Vendedor'


    for cell in 'ABC':
        hoja_salida[cell +'1'].fill = blackFill
        hoja_salida[cell +'1'].font = Font(color=WHITE, bold=True, size=20)
        hoja_salida[cell +'5'].fill = blackFill
        hoja_salida[cell +'5'].font = Font(color=WHITE, bold=True, size=20)

    for cell in 'ABCD':
        hoja_salida1[cell +'1'].fill = blackFill
        hoja_salida1[cell +'1'].font = Font(color=WHITE, bold=True, size=20)


    dim = {'A': 40, 'B': 35, 'C': 20, 'D': 40}
    for k, v in dim.items():
        hoja_salida.column_dimensions[k].width = v
        hoja_salida1.column_dimensions[k].width = v

# Escribir Totales
    row = 2
    for i in list_total_reporte:
        hoja_salida['A' + str(row)] = i[0]
        hoja_salida['B' + str(row)] = i[1]
        row += 1

#Escribir Totales por dia
    row1 = 6
    for i in list_prendas_facturas_dia:
        hoja_salida['A' + str(row1)] = i[0]
        hoja_salida['B' + str(row1)] = i[1]
        hoja_salida['C' + str(row1)] = i[2]
        row1 += 1

#Escribir Totales por dia por vendedora
    row2 = 2
    for i in list_por_vendedora:
        hoja_salida1['A' + str(row2)] = i[0]
        hoja_salida1['B' + str(row2)] = i[2]
        hoja_salida1['C' + str(row2)] = i[3]
        hoja_salida1['D' + str(row2)] = i[1]
        row2 += 1
    fecha = datetime.date.today()
    excel_salida.save("/Users/mpatinob/Dropbox/Reportes_LP/Reportes Procesados/"+ str(fecha) + "_Procesado_" + str(file))


path = '/Users/mpatinob/Dropbox/Reportes_LP/Reportes SAP'
files = [f for f in listdir(path) if isfile(join(path, f))]

for file in files:
    if file != '.DS_Store':
        try:
            wb = openpyxl.load_workbook('/Users/mpatinob/Dropbox/Reportes_LP/Reportes SAP/' + str(file))
            ws = wb.worksheets[0]
            lista_inicial = lectura_reporte(ws)
            db, vendedor = crear_db(lista_inicial)
            list_total_reporte, list_prendas_facturas_dia, list_por_vendedora = procesar_db(db)
            print_report(list_total_reporte, list_prendas_facturas_dia, list_por_vendedora, file)
        except KeyError:
            print "Archivo Corrupto o No soportado"
